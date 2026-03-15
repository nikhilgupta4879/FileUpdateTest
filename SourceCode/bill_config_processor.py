"""
bill_config_processor.py
------------------------
Processes BillUpdateConfig03152026.xlsx to:
  - Rename/archive current bill image files
  - Copy new bill image files into place
  - Generate SQL update scripts for database remediation

Usage:
    python bill_config_processor.py [--base-path <path>] [--config-path <path>]

Defaults (relative to project root):
    --base-path    TestFiles
    --config-path  Config/BillUpdateConfig03152026.xlsx
"""

import os
import shutil
import argparse
import logging
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Logging setup
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
CONFIG_FILENAME     = "BillUpdateConfig03152026.xlsx"
SCRIPT_SUBDIR       = "scripts"
SQL_FILE_GUID_RESET = "ResetCurrGuidToNull.txt"
SQL_FILE_TASK_RESET = "UpdateTaskToPushImageAgain.txt"

# Excel column names (row 1 is the header row, row 2 is blank, data starts row 3)
COL_SEQ_NO          = "S. No"
COL_COMPANY_CODE    = "CompanyCode"
COL_BILL_IMG_PATH   = "BillImagePath"
COL_BILL_IMG_NAME   = "BillImageName"
COL_CURRENT_GUID    = "CurrentGUID"
COL_IMAGE_ID        = "ImageID"
COL_TASK_ID         = "taskId"
COL_NEW_IMG_PATH    = "NewBillImagePath"
COL_NEW_IMG_NAME    = "NewBillImageName"


# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------

def normalize_path(raw_path: str) -> str:
    """
    Convert a Windows-style path fragment (leading backslash, backslash
    separators) to a POSIX-compatible relative path without a leading slash.

    Examples:
        \\aic\\bill\\03142026\\  ->  aic/bill/03142026
        \\newImages\\            ->  newImages
    """
    # Replace backslashes with forward slashes, then strip surrounding slashes
    return raw_path.replace("\\", "/").strip("/")


def resolve_path(base: Path, path_fragment: str) -> Path:
    """Return an absolute Path by joining *base* with the normalised fragment."""
    return base / normalize_path(path_fragment)


def append_sql(script_path: Path, statement: str) -> None:
    """Append a SQL statement (terminated with ';') to *script_path*."""
    script_path.parent.mkdir(parents=True, exist_ok=True)
    with script_path.open("a", encoding="utf-8") as fh:
        fh.write(statement + "\n")


def load_config(config_file: Path) -> list[dict]:
    """
    Load the Excel config file and return a list of row-dicts.
    Row 1  = header, Row 2 = blank (skipped), data starts at Row 3.
    Only rows where S. No is a non-empty integer are processed.
    """
    wb = openpyxl.load_workbook(config_file)
    ws = wb.active

    # Build header map: column-name -> 0-based index
    header_row = [cell.value for cell in ws[1]]
    headers = {name: idx for idx, name in enumerate(header_row) if name}

    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        seq = row[headers[COL_SEQ_NO]]
        if seq is None:
            continue  # skip blank rows
        record = {col: row[idx] for col, idx in headers.items()}
        rows.append(record)

    log.info("Loaded %d data rows from %s", len(rows), config_file.name)
    return rows


# ---------------------------------------------------------------------------
# Per-row processing
# ---------------------------------------------------------------------------

def process_row(row: dict, base_path: Path, script_dir: Path) -> None:
    """Execute all processing steps for a single config row."""

    seq          = row[COL_SEQ_NO]
    company_code = str(row[COL_COMPANY_CODE]).strip()
    bill_img_dir = resolve_path(base_path, str(row[COL_BILL_IMG_PATH]).strip())
    bill_img_name = str(row[COL_BILL_IMG_NAME]).strip()
    current_guid = str(row[COL_CURRENT_GUID]).strip()
    image_id     = str(row[COL_IMAGE_ID]).strip()
    task_id      = str(row[COL_TASK_ID]).strip()
    new_img_dir  = resolve_path(base_path, str(row[COL_NEW_IMG_PATH]).strip())
    new_img_name = str(row[COL_NEW_IMG_NAME]).strip()

    log.info("--- Processing row %s | %s | %s ---", seq, company_code, bill_img_name)

    # ------------------------------------------------------------------
    # Step 3 – Rename the current bill image to <name>_toberenamed.<ext>
    # ------------------------------------------------------------------
    current_file = bill_img_dir / bill_img_name
    stem, suffix = os.path.splitext(bill_img_name)
    archived_name = f"{stem}_toberenamed{suffix}"
    archived_file = bill_img_dir / archived_name

    if not current_file.exists():
        log.warning("  [Step 3] Source file not found, skipping rename: %s", current_file)
    else:
        current_file.rename(archived_file)
        log.info("  [Step 3] Renamed  %s  ->  %s", current_file.name, archived_file.name)

    # ------------------------------------------------------------------
    # Step 4 – Copy new image file into the bill image directory
    # ------------------------------------------------------------------
    new_source_file = new_img_dir / new_img_name
    new_dest_file   = bill_img_dir / new_img_name

    if not new_source_file.exists():
        log.warning("  [Step 4] New source file not found, skipping copy: %s", new_source_file)
    else:
        bill_img_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(new_source_file, new_dest_file)
        log.info("  [Step 4] Copied   %s  ->  %s", new_source_file, new_dest_file)

    # ------------------------------------------------------------------
    # Step 5 – Rename copied file from NewBillImageName to BillImageName
    # ------------------------------------------------------------------
    renamed_dest = bill_img_dir / bill_img_name

    if not new_dest_file.exists():
        log.warning("  [Step 5] Copied file not found, skipping rename: %s", new_dest_file)
    else:
        new_dest_file.rename(renamed_dest)
        log.info("  [Step 5] Renamed  %s  ->  %s", new_dest_file.name, renamed_dest.name)

    # ------------------------------------------------------------------
    # Steps 6 & 7 – SQL: reset externalRefGUID to NULL
    # ------------------------------------------------------------------
    sql_guid_reset = (
        f"update Integration.bill.BillImage "
        f"set externalRefGUID = null "
        f"where billImageId = '{image_id}' "
        f"and externalRefGUID='{current_guid}';"
    )
    guid_script = script_dir / SQL_FILE_GUID_RESET
    append_sql(guid_script, sql_guid_reset)
    log.info("  [Step 6-7] Appended GUID-reset SQL to %s", guid_script.name)

    # ------------------------------------------------------------------
    # Steps 8 & 9 – SQL: reset Task to Ready
    # ------------------------------------------------------------------
    sql_task_reset = (
        f"update Integration.dbo.Task "
        f"set taskStatus='Ready',processCounter=0,batchId=null "
        f"where taskId = '{task_id}' "
        f"and companyCode='{company_code}';"
    )
    task_script = script_dir / SQL_FILE_TASK_RESET
    append_sql(task_script, sql_task_reset)
    log.info("  [Step 8-9] Appended Task-reset SQL to %s", task_script.name)


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Process BillUpdateConfig and generate file + SQL artefacts."
    )
    parser.add_argument(
        "--base-path",
        default="TestFiles",
        help="Root directory for bill image files (default: TestFiles)",
    )
    parser.add_argument(
        "--config-path",
        default=os.path.join("Config", CONFIG_FILENAME),
        help=f"Path to the Excel config file (default: Config/{CONFIG_FILENAME})",
    )
    args = parser.parse_args()

    # Resolve paths relative to the script's own directory so the script
    # works correctly regardless of the caller's working directory.
    script_dir_root = Path(__file__).resolve().parent.parent  # project root
    base_path   = (script_dir_root / args.base_path).resolve()
    config_file = (script_dir_root / args.config_path).resolve()
    script_out  = base_path / SCRIPT_SUBDIR

    log.info("Project root : %s", script_dir_root)
    log.info("Config file  : %s", config_file)
    log.info("Base path    : %s", base_path)
    log.info("Script output: %s", script_out)

    if not config_file.exists():
        log.error("Config file not found: %s", config_file)
        raise SystemExit(1)

    rows = load_config(config_file)

    success_count = 0
    error_count   = 0
    for row in rows:
        try:
            process_row(row, base_path, script_out)
            success_count += 1
        except Exception as exc:          # pylint: disable=broad-except
            log.error("  Row %s failed: %s", row.get(COL_SEQ_NO), exc)
            error_count += 1

    log.info(
        "=== Completed: %d row(s) processed, %d error(s) ===",
        success_count,
        error_count,
    )


if __name__ == "__main__":
    main()
